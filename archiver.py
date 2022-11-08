from zipfile import ZipFile
import csv
import io
from openpyxl import load_workbook
import PyPDF2

with ZipFile('resources/test_zip.zip', 'w') as zip_file:
    zip_file.write('test_files2/test_csv.csv', 'csv_test.csv')
    zip_file.write('test_files2/test_pdf.pdf', 'pdf_test.pdf')
    zip_file.write('test_files2/test_xlsx.xlsx', 'xlsx_test.xlsx')

with ZipFile('resources/test_zip.zip') as csvfile:
    with csvfile.open('csv_test.csv', mode='r') as file:
        table = csv.reader(io.TextIOWrapper(file, 'utf-8'))
        for line_no, line in enumerate(table, 1):
            if line_no == 2:
                assert 'H01' in line[5]

with ZipFile('resources/test_zip.zip') as pdffile:
    pdf = pdffile.read('pdf_test.pdf')
    pdf_reader = PyPDF2.PdfReader(io.BytesIO(pdf))
    page = pdf_reader.pages[0]
    text = page.extract_text()
    assert '1. Что необходимо тестировать?' in text

with ZipFile('resources/test_zip.zip') as xlsxfile:
    xlsx = xlsxfile.read('xlsx_test.xlsx')
    workbook = load_workbook(io.BytesIO(xlsx))
    sheet = workbook.active
    text = sheet.cell(row=15, column=6).value
    assert '"Лицей"Физико-техническая школа"' in text
